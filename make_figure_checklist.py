from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(r"D:\laptop file sync\#Rutgers PhD project\altermagnetic\new general path plotting band method\structure\all paper structures")
OUT = ROOT / "figure_checklist_cubic_to_monoclinic.xlsx"


CASES = [
    (1, 42, "m-3m", "cP (Pm-3m)", "CrO SG223", r"7-cubic\Pm-3m-B-i", "01_c-Pm3m.png"),
    (2, 43, "m-3m", "cF (Fm-3m)", "Mn8Te96 SG226", r"7-cubic\Fm-3m-B-i", "02_c-Fm3m.png"),
    (3, 44, "m-3m", "cI (Im-3m)", "Mn16Te16 SG230", r"7-cubic\Im-3m-B-i", "03_c-Im3m.png"),
    (4, 1, "6/mmm", "hP (SG 177-194)", "P63cm", r"6-hexagonal\6-mmm\2221-B-g", "04_h-6mmm-2221.png"),
    (5, 2, "6/mmm", "hP (SG 177-194)", "GdAuGe", r"6-hexagonal\6-mmm\2212-B-g", "05_2212-B-g.png"),
    (6, 3, "6/mmm", "hP (SG 177-194)", "MnO SG193", r"6-hexagonal\6-mmm\1122-P-i", "06_h-6mmm-1122.png"),
    (7, 4, "6/m", "hP (SG 168-176)", "P6_3 compound", r"6-hexagonal\6-m-B-g", "07_h-6m.png"),
    (8, 39, "-3m", "hP (SG 165, P-3c1)", "MnCu2P", r"5-trigonal\Primitive", "08_tri-Primitive.png"),
    (9, 40, "-3m", "RHL (R3c)", "BiFeO3", r"5-trigonal\rhombohedral\case-1", "09_tri-case1.png"),
    (10, 41, "-3m", "RHL (R-3c)", "CrF3", r"5-trigonal\rhombohedral\case-2", "10_tri-case2.png"),
    (11, 5, "4/mmm", "tP1 (P4/mmm)", "MnF2 (0.15)", r"4-tetragonal\4-mmm\P4-mmm\2121-P-d", "11_t-P4mmm-2121.png"),
    (12, 6, "4/mmm", "tP1 (P4/mmm)", "KV2Se2O (1.875)", r"4-tetragonal\4-mmm\P4-mmm\2112-P-d", "12_t-P4mmm-2112.png"),
    (13, 7, "4/mmm", "tP1 (P4/mmm)", "MnAlPt2", r"4-tetragonal\4-mmm\P4-mmm\1122-P-g", "13_t-P4mmm-1122.png"),
    (14, 8, "4/mmm", "tI1 c<a (I4/mmm)", "CaFe4Al8 (0.236)", r"4-tetragonal\4-mmm\I4-mmm-c-lt-a\2121-P-d", "14_t-I4mmmclt-2121.png"),
    (15, 9, "4/mmm", "tI1 c<a (I4/mmm)", "BaZnSb2", r"4-tetragonal\4-mmm\I4-mmm-c-lt-a\2112-P-d", "15_t-I4mmmclt-2112.png"),
    (16, 10, "4/mmm", "tI1 c<a (I4/mmm)", "KMnF3", r"4-tetragonal\4-mmm\I4-mmm-c-lt-a\1122-P-g", "16_t-I4mmmclt-1122.png"),
    (17, 11, "4/mmm", "tI2 c>a (I4/mmm)", "CaFe4Al8 c>a", r"4-tetragonal\4-mmm\I4-mmm-c-gt-a\2121-P-d", "17_t-I4mmmcgt-2121.png"),
    (18, 12, "4/mmm", "tI2 c>a (I4/mmm)", "BaZnSb2 c>a", r"4-tetragonal\4-mmm\I4-mmm-c-gt-a\2112-P-d", "18_t-I4mmmcgt-2112.png"),
    (19, 13, "4/mmm", "tI2 c>a (I4/mmm)", "KMnF3 c>a", r"4-tetragonal\4-mmm\I4-mmm-c-gt-a\1122-P-g", "19_t-I4mmmcgt-1122.png"),
    (20, 14, "4/m", "tP (P4/m)", "P4/m compound", r"4-tetragonal\4-m\P4-m", "20_t-P4m.png"),
    (21, 15, "4/m", "tI c<a (I4/m)", "I4/m compound", r"4-tetragonal\4-m\I4-m-c-lt-a", "21_t-I4mclt.png"),
    (22, 16, "4/m", "tI c>a (I4/m)", "DyNbO4", r"4-tetragonal\4-m\I4-m-c-gt-a", "22_t-I4mcgt.png"),
    (23, 17, "mmm", "oP1 (Pmmm)", "MnTe", r"3-orthorombic\Pmmm\221-P-d", "23_o-Pmmm-221.png"),
    (24, 18, "mmm", "oP1 (Pmmm)", "ScCrO3 (0.307)", r"3-orthorombic\Pmmm\212-B-d", "24_o-Pmmm-212.png"),
    (25, 19, "mmm", "oP1 (Pmmm)", "MnTe B3g", r"3-orthorombic\Pmmm\122-B-d", "25_o-Pmmm-122.png"),
    (26, 20, "mmm", "oI (Immm)", "YBaMn2O5 (0.98)", r"3-orthorombic\Immm\221-P-d", "26_o-Immm-221.png"),
    (27, 21, "mmm", "oI (Immm)", "ZnFeF5(H2O)2 (0.575)", r"3-orthorombic\Immm\212-B-d", "27_o-Immm-212.png"),
    (28, 22, "mmm", "oI (Immm)", "MnO primitive", r"3-orthorombic\Immm\122-B-d", "28_o-Immm-122.png"),
    (29, 23, "mmm", "oC (Cmmm)", "Sr4Fe4O11 (0.402)", r"3-orthorombic\Cmmm\221-P-d", "29_o-Cmmm-221.png"),
    (30, 24, "mmm", "oC (Cmmm)", "MnO B2g", r"3-orthorombic\Cmmm\212-B-d", "30_o-Cmmm-212.png"),
    (31, 25, "mmm", "oC (Cmmm)", "MnO B3g", r"3-orthorombic\Cmmm\122-B-d", "31_o-Cmmm-122.png"),
    (32, 26, "mmm", "oF1 (Fmmm)", "generated oF1", r"3-orthorombic\Fmmm\oF1\221-P-d", "32_o-Fmmm-oF1-221.png"),
    (33, 27, "mmm", "oF1 (Fmmm)", "generated oF1", r"3-orthorombic\Fmmm\oF1\212-B-d", "33_o-Fmmm-oF1-212.png"),
    (34, 28, "mmm", "oF1 (Fmmm)", "generated oF1", r"3-orthorombic\Fmmm\oF1\122-B-d", "34_o-Fmmm-oF1-122.png"),
    (35, 29, "mmm", "oF2 (Fdd2)", "generated oF2", r"3-orthorombic\Fmmm\oF2\221-P-d", "35_o-Fmmm-oF2-221.png"),
    (36, 30, "mmm", "oF2 (Fdd2)", "generated oF2", r"3-orthorombic\Fmmm\oF2\122-B-d", "36_o-Fmmm-oF2-122.png"),
    (37, 31, "mmm", "oF2 (Fdd2)", "generated oF2", r"3-orthorombic\Fmmm\oF2\212-B-d", "37_o-Fmmm-oF2-212.png"),
    (38, 32, "mmm", "oF3 (Fmmm)", "generated oF3", r"3-orthorombic\Fmmm\oF3\221-P-d", "38_o-Fmmm-oF3-221.png"),
    (39, 33, "mmm", "oF3 (Fmmm)", "generated oF3", r"3-orthorombic\Fmmm\oF3\212-B-d", "39_o-Fmmm-oF3-212.png"),
    (40, 34, "mmm", "oF3 (Fmmm)", "generated oF3", r"3-orthorombic\Fmmm\oF3\122-B-d", "40_o-Fmmm-oF3-122.png"),
    (41, 35, "2/m", "mP (P2/m)", "generated mP", r"2-monoclinic\P2-m-B-d", "41_m-P2m.png"),
    (42, 36, "2/m", "mC1 (C2/m)", "generated mC1", r"2-monoclinic\C-2m-B-d\mC1", "42_m-C2m-mC1.png"),
    (43, 37, "2/m", "mC2 (C2/m)", "generated mC2", r"2-monoclinic\C-2m-B-d\mC2", "43_m-C2m-mC2.png"),
    (44, 38, "2/m", "mC3 (C2/m)", "generated mC3", r"2-monoclinic\C-2m-B-d\mC3", "44_m-C2m-mC3.png"),
]


def has_pattern(folder: Path, token: str) -> str:
    return "yes" if any(folder.glob(f"*_{token}_*.png")) else ""


wb = Workbook()
ws = wb.active
ws.title = "Figure Checklist"

headers = [
    "Check order",
    "Markdown case #",
    "Laue",
    "Lattice type",
    "Example material",
    "Structure folder",
    "Fig 1 IBZ checked",
    "Fig 2 spinflip checked",
    "Fig 3 spinBZ checked",
    "Fig 4 top checked",
    "Band figure checked",
    "Geometry/path OK?",
    "Label positions OK?",
    "Spin coloring OK?",
    "Needs rerun?",
    "Notes",
    "IBZ PNG exists",
    "Spinflip PNG exists",
    "SpinBZ PNG exists",
    "Top PNG exists",
    "Band figure exists",
    "Band figure file",
]
ws.append(headers)

for order, md_case, laue, lattice, material, rel_folder, band_file in CASES:
    folder = ROOT / rel_folder
    band_path = ROOT / "band_figures" / band_file
    ws.append([
        order,
        md_case,
        laue,
        lattice,
        material,
        rel_folder,
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        "",
        has_pattern(folder, "ibz"),
        has_pattern(folder, "spinflip"),
        has_pattern(folder, "spinbz"),
        has_pattern(folder, "spinbz_top"),
        "yes" if band_path.exists() else "MISSING",
        str(band_path),
    ])

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(CASES) + 1}"

header_fill = PatternFill("solid", fgColor="1F4E78")
missing_fill = PatternFill("solid", fgColor="FFC7CE")
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    if row[20].value == "MISSING":
        for cell in row:
            cell.fill = missing_fill

widths = {
    "A": 11,
    "B": 14,
    "C": 10,
    "D": 22,
    "E": 24,
    "F": 48,
    "G": 16,
    "H": 20,
    "I": 18,
    "J": 16,
    "K": 18,
    "L": 16,
    "M": 18,
    "N": 18,
    "O": 14,
    "P": 36,
    "Q": 14,
    "R": 17,
    "S": 15,
    "T": 14,
    "U": 16,
    "V": 58,
}
for col, width in widths.items():
    ws.column_dimensions[col].width = width

table_ref = f"A1:{get_column_letter(len(headers))}{len(CASES) + 1}"
tab = Table(displayName="FigureChecklist", ref=table_ref)
tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
ws.add_table(tab)

summary = wb.create_sheet("Summary")
summary.append(["Item", "Value"])
summary.append(["Total checklist rows", len(CASES)])
summary.append(["Band figures present", sum((ROOT / "band_figures" / c[-1]).exists() for c in CASES)])
summary.append(["Band figures missing", ", ".join(c[-1] for c in CASES if not (ROOT / "band_figures" / c[-1]).exists())])
summary.append(["Order", "Cubic -> hexagonal -> trigonal -> tetragonal -> orthorhombic -> monoclinic"])
summary.append(["Source", "CLAUDE.md case table plus actual all paper structures/band_figures folders"])
for row in summary.iter_rows():
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
summary["A1"].font = Font(bold=True)
summary["B1"].font = Font(bold=True)
summary.column_dimensions["A"].width = 24
summary.column_dimensions["B"].width = 90

wb.save(OUT)
print(OUT)
